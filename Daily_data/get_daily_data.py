import pandas as pd
from bs4 import BeautifulSoup as soup
import os
from urllib.request import Request, urlopen
import numpy as np
from openpyxl import load_workbook
import time
from random import random


def get_number(number):
    negative = False
    if "B" in number:
        return float(number.replace("B", "")) * 1000000000
    if "-" in number:
        number = number.replace("-", "")
        negative = True
    number = number.replace(",", "").replace("%", "").replace("$", "")
    if negative:
        return -float(number)
    else:
        return float(number)


def get_ticker_list():
    path = os.path.dirname(os.path.abspath(__file__))
    df = pd.read_csv(path + "/ticker_list.csv")
    return df


def get_fundamentals(symbol):
    url = "http://finviz.com/quote.ashx?t=" + symbol.lower()
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    webpage = urlopen(req).read()
    html = soup(webpage, "html.parser")

    fundamentals = pd.read_html(str(html), attrs={"class": "snapshot-table2"})[0]

    fundamentals.columns = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11"]
    colOne = []
    colLength = len(fundamentals)
    for k in np.arange(0, colLength, 2):
        colOne.append(fundamentals[f"{k}"])
    attrs = pd.concat(colOne, ignore_index=True)

    colTwo = []
    colLength = len(fundamentals)
    for k in np.arange(1, colLength, 2):
        colTwo.append(fundamentals[f"{k}"])
    vals = pd.concat(colTwo, ignore_index=True)

    fundamentals = pd.DataFrame()
    fundamentals["Attributes"] = attrs
    fundamentals["Values"] = vals
    fundamentals = fundamentals.set_index("Attributes")
    return fundamentals


def get_each_price_data(df):
    df[["MarketCap", "UpDown", "Price", "52High", "RSI"]] = 0
    for i in range(len(df)):
        fu = get_fundamentals(df.loc[i, "Ticker"])
        df.loc[i, "MarketCap"] = get_number(fu.loc["Market Cap", "Values"])
        df.loc[i, "UpDown"] = get_number(fu.loc["Change", "Values"])
        df.loc[i, "Price"] = get_number(fu.loc["Price", "Values"])
        df.loc[i, "52High"] = get_number(fu.loc["52W High", "Values"])
        df.loc[i, "RSI"] = get_number(fu.loc["RSI (14)", "Values"])
        rand_value = random() * 2 + 1
        time.sleep(rand_value)
    return df


def sort_by_marketcap(df):
    df.sort_values("MarketCap", ascending=False, inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


def write_to_excel(df):
    path = os.path.dirname(os.path.abspath(__file__))
    excel = load_workbook(path + "/daily_data_reference.xlsx")
    sheet = excel.active
    columns = df.columns
    for i in range(len(df)):
        for j in range(len(columns)):
            sheet.cell(i + 4, j + 2, df.loc[i, columns[j]])
    excel.save(path + "/daily_data.xlsx")
    excel.close()
    df["Ticker"].to_csv(path + "/ticker_list.csv", index=False)


def main():
    df = get_ticker_list()
    df = get_each_price_data(df)
    df = sort_by_marketcap(df)
    write_to_excel(df)


if __name__ == "__main__":
    main()
